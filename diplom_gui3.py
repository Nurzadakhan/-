import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from PIL import Image, ImageDraw, ImageFont, ImageTk
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
import os
import sys
from openpyxl import load_workbook
from openpyxl import Workbook

# ------------------------------------------------------------
#  ФУНКЦИЯ ГЕНЕРАЦИИ (читает Excel сама, без pandas)
# ------------------------------------------------------------
def generate_diplomas(excel_path, is_duplicate, output_path, progress_callback=None):
    # Определяем пути к ресурсам (для exe или для скрипта)
    if getattr(sys, 'frozen', False):
        base_path = sys._MEIPASS
    else:
        base_path = os.path.dirname(os.path.abspath(__file__))

    template_path = os.path.join(base_path, "template.png")
    font_path = os.path.join(base_path, "times.ttf")
    font_italic_path = os.path.join(base_path, "timesi.ttf")

    # Проверка наличия шаблона и шрифтов
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Шаблон не найден: {template_path}")
    if not os.path.exists(font_path):
        raise FileNotFoundError(f"Шрифт не найден: {font_path}")

    # --- Чтение Excel через openpyxl ---
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active

    # Заголовки (первая строка)
    headers = [cell.value for cell in ws[1]]
    # Данные (со второй строки)
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data.append(dict(zip(headers, row)))

    if not data:
        raise ValueError("Excel-файл не содержит строк с данными (после заголовков).")

    # Проверка наличия всех необходимых колонок
    required_columns = [
        'fio_kz', 'fio_ru', 'o_kz', 'o_ru', 'year_start', 'year_end',
        'specialty_kz', 'specialty_ru', 'qualification_kz', 'qualification_ru',
        'city_kz', 'city_ru', 'form_kz', 'form_ru', 'signature_1', 'signature_2',
        'year_kz', 'day_kz', 'month_kz', 'month_ru', 'month_ru2', 'reg_number',
        'institution_kz', 'institution_ru'
    ]
    missing = [col for col in required_columns if col not in headers]
    if missing:
        raise ValueError(f"В файле отсутствуют столбцы: {', '.join(missing)}")

    template_pil = Image.open(template_path).convert("RGB")
    TEMPLATE_WIDTH, TEMPLATE_HEIGHT = template_pil.size
    temp_images = []

    for idx, row in enumerate(data):
        if progress_callback:
            progress_callback(idx, len(data))

        img = template_pil.copy()
        draw = ImageDraw.Draw(img)
        font = ImageFont.truetype(font_path, 35)

        # -------------------------- ВСЕ ВАШИ draw.text() --------------------------
        draw.text((420, 700), row['fio_kz'], font=font, fill="black")
        draw.text((208, 750), row['o_kz'], font=font, fill="black")
        draw.text((260, 790), str(row['year_start']), font=font, fill="black")
        draw.text((297, 878), str(row['year_end']), font=font, fill="black")
        draw.text((200, 1028), row['specialty_kz'], font=font, fill="black")
        draw.text((220, 1280), row['qualification_kz'], font=font, fill="black")

        draw.text((1861, 700), row['fio_ru'], font=font, fill="black")
        draw.text((1444, 744), row['o_ru'], font=font, fill="black")
        draw.text((1732, 788), str(row['year_start']), font=font, fill="black")
        draw.text((1487, 943), str(row['year_end']), font=font, fill="black")
        draw.text((1430, 1080), row['specialty_ru'], font=font, fill="black")
        draw.text((1450, 1308), row['qualification_ru'], font=font, fill="black")

        draw.text((480, 1580), row['city_kz'], font=font, fill="black")
        draw.text((1848, 1576), row['city_ru'], font=font, fill="black")
        draw.text((450, 1117), row['form_kz'], font=font, fill="black")
        draw.text((1721, 1130), row['form_ru'], font=font, fill="black")
        draw.text((262, 1620), str(row['year_kz']), font=font, fill="black")
        draw.text((560, 1620), str(row['day_kz']), font=font, fill="black")
        draw.text((700, 1620), row['month_kz'], font=font, fill="black")
        draw.text((864, 1168), str(row['year_kz']), font=font, fill="black")
        draw.text((342, 1216), str(row['day_kz']), font=font, fill="black")
        draw.text((522, 1216), row['month_kz'], font=font, fill="black")
        draw.text((1984, 1625), str(row['year_kz']), font=font, fill="black")
        draw.text((1549, 1625), str(row['day_kz']), font=font, fill="black")
        draw.text((1715, 1625), row['month_ru'], font=font, fill="black")
        draw.text((2209, 1196), str(row['year_kz']), font=font, fill="black")
        draw.text((1570, 1234), str(row['day_kz']), font=font, fill="black")
        draw.text((1437, 1234), row['month_ru2'], font=font, fill="black")
        draw.text((682, 1670), str(row['reg_number']), font=font, fill="black")
        draw.text((2000, 1670), str(row['reg_number']), font=font, fill="black")
        draw.text((200, 838), row['institution_kz'], font=font, fill="black")
        draw.text((204, 934), row['institution_kz'], font=font, fill="black")
        draw.text((1440, 880), row['institution_ru'], font=font, fill="black")
        draw.text((1440, 987), row['institution_ru'], font=font, fill="black")
        draw.text((705, 1395), row['signature_1'], font=font, fill="black")
        draw.text((1985, 1395), row['signature_1'], font=font, fill="black")
        draw.text((705, 1465), row['signature_2'], font=font, fill="black")
        draw.text((1985, 1465), row['signature_2'], font=font, fill="black")
        # -------------------------------------------------------------

        # Дубликат (курсив из встроенного файла)
        if is_duplicate:
            font_dup = ImageFont.truetype(font_italic_path, 35)
            dup_text = "Дубликат"
            x = TEMPLATE_WIDTH - 500
            y = 25
            draw.text((x, y), dup_text, font=font_dup, fill="black")

        temp_path = f"temp_{idx}.jpg"
        img.save(temp_path, "JPEG", quality=95)
        temp_images.append(temp_path)

    # Сборка PDF
    c = canvas.Canvas(output_path, pagesize=A4)
    pw, ph = A4
    for img_path in temp_images:
        c.drawImage(img_path, 0, 0, width=pw, height=ph)
        c.showPage()
    c.save()

    # Чистка временных файлов
    for path in temp_images:
        os.remove(path)

    return output_path


# ------------------------------------------------------------
#  GUI НА TKINTER
# ------------------------------------------------------------
class DiplomaApp:
    def __init__(self, root):
        self.root = root
        root.title("Генератор дипломов")
        root.geometry("600x450")
        root.resizable(False, False)

        # Переменные
        self.excel_path = tk.StringVar()
        self.is_duplicate = tk.BooleanVar()

        # Загрузка логотипа (вшитый)
        if getattr(sys, 'frozen', False):
            base_path = sys._MEIPASS
        else:
            base_path = os.path.dirname(os.path.abspath(__file__))

        logo_path = os.path.join(base_path, "logo.png")
        try:
            logo_img = Image.open(logo_path)
            logo_img = logo_img.resize((100, 100), Image.Resampling.LANCZOS)
            self.logo = ImageTk.PhotoImage(logo_img)
            logo_label = tk.Label(root, image=self.logo)
            logo_label.pack(pady=(12, 0))
        except Exception as e:
            print(f"Не удалось загрузить логотип: {e}")

        tk.Label(root, text="Генератор дипломов для ТИПО", font=("Arial", 12, "bold")).pack()
        tk.Label(root, text="Заполните данные и нажмите кнопку").pack(pady=(0, 10))

        # Виджеты
        tk.Label(root, text="1. Выберите Excel-файл с данными:").pack(pady=(15, 5))
        frm = tk.Frame(root)
        frm.pack(fill='x', padx=20)
        tk.Entry(frm, textvariable=self.excel_path, width=40).pack(side='left', fill='x', expand=True)
        tk.Button(frm, text="Обзор", command=self.select_excel).pack(side='right', padx=(5, 0))

        tk.Checkbutton(root, text="Это дубликат (добавить надпись)", variable=self.is_duplicate).pack(pady=15)

        self.btn_generate = tk.Button(root, text="Создать дипломы!", command=self.run_generation,
                                      bg="#4CAF50", fg="white", font=("Arial", 10, "bold"))
        self.btn_generate.pack(pady=10)

        self.status = tk.Label(root, text="Готов к работе", fg="gray")
        self.status.pack(pady=(10, 0))

        self.btn_template = tk.Button(root, text="📎 Скачать шаблон Excel", command=self.download_template)
        self.btn_template.pack(pady=5)

        self.progress = ttk.Progressbar(root, orient='horizontal', length=400, mode='determinate')
        self.progress.pack(pady=10)

    def select_excel(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if path:
            self.excel_path.set(path)
            self.status.config(text=f"Выбран файл: {os.path.basename(path)}", fg="blue")

    def update_progress(self, current, total):
        self.progress['maximum'] = total
        self.progress['value'] = current + 1
        self.root.update_idletasks()

    def run_generation(self):
        excel_file = self.excel_path.get()
        if not excel_file or not os.path.exists(excel_file):
            messagebox.showerror("Ошибка", "Пожалуйста, выберите существующий Excel-файл.")
            return

        # Диалог выбора места сохранения PDF
        output_pdf_path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf")],
            initialfile="diplomas.pdf",
            title="Сохранить дипломы как..."
        )
        if not output_pdf_path:
            return

        self.btn_generate.config(state='disabled', text="Генерация...")
        self.status.config(text="Идёт создание дипломов...", fg="orange")
        self.progress['value'] = 0

        try:
            generated_file = generate_diplomas(
                excel_path=excel_file,
                is_duplicate=self.is_duplicate.get(),
                output_path=output_pdf_path,
                progress_callback=self.update_progress
            )
            self.status.config(text=f"✅ Готово! Файл: {os.path.basename(generated_file)}", fg="green")
            messagebox.showinfo("Успех", f"Дипломы сохранены:\n{generated_file}")
        except Exception as e:
            self.status.config(text=f"❌ Ошибка: {str(e)}", fg="red")
            messagebox.showerror("Ошибка", str(e))
        finally:
            self.btn_generate.config(state='normal', text="Создать дипломы!")
            self.progress['value'] = 0

    def download_template(self):
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile="template_students.xlsx"
        )
        if not file_path:
            return

        columns = [
            'fio_kz', 'fio_ru', 'o_kz', 'o_ru', 'year_start', 'year_end',
            'specialty_kz', 'specialty_ru', 'qualification_kz', 'qualification_ru',
            'city_kz', 'city_ru', 'form_kz', 'form_ru', 'signature_1', 'signature_2',
            'year_kz', 'day_kz', 'month_kz', 'month_ru', 'month_ru2', 'reg_number',
            'institution_kz', 'institution_ru'
        ]
        # Создаём шаблон через openpyxl (без pandas)
        wb = Workbook()
        ws = wb.active
        ws.append(columns)  # заголовки
        # можно добавить строку-пример (раскомментировать при желании)
        # example_row = ["Иванов Иван", "Ivanov Ivan", ...]
        # ws.append(example_row)
        try:
            wb.save(file_path)
            messagebox.showinfo("Успех", f"Шаблон сохранён:\n{file_path}")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось сохранить шаблон:\n{e}")


# ------------------------------------------------------------
if __name__ == "__main__":
    root = tk.Tk()
    app = DiplomaApp(root)
    root.mainloop()